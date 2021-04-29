import os
import time

import openpyxl
import openpyxl.styles
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

from backend.driver import clock
from backend.driver import alert as altctl
from backend.settings import SYS_CONF

STATUS_MAPPING = {
    "opening": "未关闭",
    "closed": "已关闭",
}

class Write(object):
    def __init__(self, excel_name, name):
        self.excel_name = "{}.xlsx".format(excel_name)
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = name

        border_style = NamedStyle(name="border_style")
        bian = Side(style="medium", color="000000")
        border = Border(top=bian, bottom=bian, left=bian, right=bian)
        border_style.border = border
        alignment = Alignment(horizontal="center", vertical="center")
        border_style.alignment = alignment

        self.border_style = border_style
        self.workbook.add_named_style(border_style)

        title_style = NamedStyle(name="title_style")
        ft = Font(name="Noto Sans CJK SC Regular", color="FFFFFF", size=11, b=False)
        fill = PatternFill("solid", fgColor="00A3FF")
        title_style.font = ft
        title_style.fill = fill
        title_style.border = border
        title_style.alignment = alignment
        self.title_style = title_style
        self.workbook.add_named_style(title_style)

    def _set_border(self, row, column):
        self.worksheet.cell(row, column).style = self.border_style

    def set_title_style(self, row, column):
        self.worksheet.cell(row, column).style = self.title_style

    def set_content_style(self, row, column):
        self._set_border(row, column)

    def time_data(self, time_stamp):
        time_dt = clock._get_csttz_dt(time_stamp)
        return clock.format_dt_readable(time_dt)

    def write_alert(self, starttime, endtime, alert_count, alert_list):
        """
        当天的数据情况
        """
        export_period = "{} ~ {}".format(
            clock.format_dt_readable(clock._get_csttz_dt(starttime)),
            clock.format_dt_readable(clock._get_csttz_dt(endtime)),
        )

        self.worksheet.append(["导出告警时段", export_period])
        self.worksheet.append([])

        # 告警数量统计
        _, all_alert_type = altctl.allowed_alert_type()

        alert_type_count = []
        for alert in all_alert_type:
            alert_type_count.append(alert_count.get(alert, 0))

        self.worksheet.append(["告警类型", "告警总量"])
        for i in range(len(all_alert_type)):
            self.worksheet.append([all_alert_type[i], alert_type_count[i]])
        self.worksheet.append([])

        # 告警详情
        self.worksheet.append(["序号", "告警类型", "设备名称", "设备编号", "告警时间", "告警地点", "当前状态"])
        index = 1
        for alert in alert_list:
            self.worksheet.append(
                [
                    index,
                    alert.get("title"),
                    alert.get("device_name"),
                    alert.get("device_id"),
                    self.time_data(alert.get("create_time")),
                    alert.get("location"),
                    STATUS_MAPPING[alert.get("status")],
                ]
            )
            index = index + 1

        # 设计样式
        self.set_title_style(1, 1)

        [self.set_title_style(3, col) for col in range(1, 3)]
        for row in range(3, 4 + len(all_alert_type)):
            self.set_title_style(row, 1)
            self.set_title_style(row, 2)

        row = 3 + len(all_alert_type) + 2
        # 序号
        [self.set_title_style(row, col) for col in range(1, 8)]
        for alert in alert_list:
            row = row + 1
            [self.set_content_style(row, col) for col in range(1, 8)]

        # 单元格宽度设计
        for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
            self.worksheet.column_dimensions[col].width = 20
        excel_dir_path = SYS_CONF["excel_dir"]
        if not os.path.exists(excel_dir_path):
            os.makedirs(excel_dir_path)
        path = os.path.join(excel_dir_path, self.excel_name)
        self.workbook.save(path)
        return {"name": self.excel_name, "path": path}
